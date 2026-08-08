"""Build the daily workbook and maintain the cumulative history workbook."""

from __future__ import annotations

import datetime as dt
import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import EXPECTED_PAIRS, Rate, sort_rates

log = logging.getLogger(__name__)

DAILY_SHEET = "Reference Rates"
META_SHEET = "Metadata"
HISTORY_SHEET = "History"
WIDE_SHEET = "Wide"

COLUMNS = [
    "Date",
    "Pair",
    "Base Currency",
    "Quote Currency",
    "Unit",
    "Rate (INR)",
    "Source",
    "Retrieved At (UTC)",
]

# FBIL publishes reference rates to four decimal places.
RATE_FORMAT = "#,##0.0000"
DATE_FORMAT = "yyyy-mm-dd"
DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def rates_to_frame(rates: list[Rate]) -> pd.DataFrame:
    frame = pd.DataFrame([r.as_row() for r in sort_rates(rates)])
    if frame.empty:
        return pd.DataFrame(columns=COLUMNS)
    return frame[COLUMNS]


# --------------------------------------------------------------------------
# Daily workbook
# --------------------------------------------------------------------------


def write_daily_workbook(
    rates: list[Rate], target_date: dt.date, output_dir: str | Path
) -> Path:
    """Write reference_rates_YYYY-MM-DD.xlsx and return its path."""
    if not rates:
        raise ValueError("Refusing to write an empty workbook")

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"reference_rates_{target_date.isoformat()}.xlsx"

    frame = rates_to_frame(rates)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # startrow=2 leaves room for the title block written below.
        frame.to_excel(writer, sheet_name=DAILY_SHEET, index=False, startrow=2)
        _metadata_frame(rates, target_date).to_excel(
            writer, sheet_name=META_SHEET, index=False
        )

    workbook = load_workbook(path)
    sheet = workbook[DAILY_SHEET]

    sheet["A1"] = f"FBIL / RBI Reference Rates — {target_date:%d %B %Y}"
    sheet["A1"].font = TITLE_FONT
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    _style_table(sheet, header_row=3, n_rows=len(frame), n_cols=len(COLUMNS))
    _add_excel_table(sheet, "DailyRates", header_row=3, n_rows=len(frame),
                     n_cols=len(COLUMNS))
    sheet.freeze_panes = "A4"

    _autosize(workbook[META_SHEET])
    workbook[META_SHEET].column_dimensions["A"].width = 26
    workbook[META_SHEET].column_dimensions["B"].width = 60

    workbook.save(path)
    log.info("Wrote %s (%d rates)", path.name, len(frame))
    return path


def _metadata_frame(rates: list[Rate], target_date: dt.date) -> pd.DataFrame:
    sources = sorted({r.source for r in rates})
    missing = [p for p in EXPECTED_PAIRS if p not in {r.pair for r in rates}]
    retrieved = max(r.retrieved_at for r in rates)
    indicative = any(r.source.endswith("-ecb") for r in rates)

    items = [
        ("Rate date", target_date.isoformat()),
        ("Pairs captured", ", ".join(r.pair for r in sort_rates(rates))),
        ("Pairs missing", ", ".join(missing) if missing else "none"),
        ("Source(s)", ", ".join(sources)),
        ("Retrieved at (UTC)", retrieved.strftime("%Y-%m-%d %H:%M:%S")),
        (
            "Provenance",
            "INDICATIVE — ECB fixing, not the official FBIL benchmark"
            if indicative
            else "Official FBIL reference rate",
        ),
        (
            "JPY convention",
            "JPY/INR is quoted per 100 yen (see the Unit column)",
        ),
    ]
    return pd.DataFrame(items, columns=["Field", "Value"])


# --------------------------------------------------------------------------
# Master history workbook
# --------------------------------------------------------------------------


def read_master(path: str | Path) -> pd.DataFrame:
    """Load an existing master workbook, tolerating absence or corruption."""
    path = Path(path)
    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame(columns=COLUMNS)
    try:
        frame = pd.read_excel(path, sheet_name=HISTORY_SHEET)
    except Exception:
        log.exception("Could not read %s; starting a fresh history", path)
        return pd.DataFrame(columns=COLUMNS)

    for column in COLUMNS:
        if column not in frame.columns:
            frame[column] = pd.NA
    return frame[COLUMNS]


def merge_into_master(
    existing: pd.DataFrame, rates: list[Rate], output_path: str | Path
) -> tuple[Path, int]:
    """Append today's rates to the history, de-duplicating on (Date, Pair).

    New rows win over old ones for the same key, so a re-run after a source
    correction overwrites rather than duplicates. Returns (path, rows_added).
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    incoming = rates_to_frame(rates)
    before = len(existing)

    # Drop empty frames before concat: pandas warns (and will change dtype
    # behaviour) when an all-NA frame takes part in the concatenation.
    parts = [f for f in (existing, incoming) if not f.empty]
    combined = (
        pd.concat(parts, ignore_index=True)
        if parts
        else pd.DataFrame(columns=COLUMNS)
    )
    combined["Date"] = pd.to_datetime(combined["Date"], errors="coerce").dt.date
    combined = combined.dropna(subset=["Date", "Pair"])
    # keep="last" => the freshly scraped row replaces any earlier one.
    combined = combined.drop_duplicates(subset=["Date", "Pair"], keep="last")
    combined = combined.sort_values(["Date", "Pair"]).reset_index(drop=True)

    wide = _wide_view(combined)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        combined.to_excel(writer, sheet_name=HISTORY_SHEET, index=False)
        wide.to_excel(writer, sheet_name=WIDE_SHEET, index=False)

    workbook = load_workbook(output_path)
    _style_table(
        workbook[HISTORY_SHEET], header_row=1, n_rows=len(combined),
        n_cols=len(COLUMNS),
    )
    _add_excel_table(
        workbook[HISTORY_SHEET], "History", header_row=1, n_rows=len(combined),
        n_cols=len(COLUMNS),
    )
    workbook[HISTORY_SHEET].freeze_panes = "A2"

    _style_table(
        workbook[WIDE_SHEET], header_row=1, n_rows=len(wide),
        n_cols=len(wide.columns),
    )
    workbook[WIDE_SHEET].freeze_panes = "B2"

    workbook.save(output_path)

    added = len(combined) - before
    log.info(
        "Master history: %d rows (%+d), %d dates covered",
        len(combined),
        added,
        combined["Date"].nunique(),
    )
    return output_path, max(added, 0)


def _wide_view(long_frame: pd.DataFrame) -> pd.DataFrame:
    """One row per date, one column per pair — the shape people actually chart."""
    if long_frame.empty:
        return pd.DataFrame(columns=["Date", *EXPECTED_PAIRS])
    wide = long_frame.pivot_table(
        index="Date", columns="Pair", values="Rate (INR)", aggfunc="last"
    ).reset_index()
    wide.columns.name = None
    ordered = ["Date"] + [p for p in EXPECTED_PAIRS if p in wide.columns]
    extras = [c for c in wide.columns if c not in ordered]
    return wide[ordered + extras]


# --------------------------------------------------------------------------
# Styling helpers
# --------------------------------------------------------------------------


def _style_table(sheet, header_row: int, n_rows: int, n_cols: int) -> None:
    for column_index in range(1, n_cols + 1):
        cell = sheet.cell(row=header_row, column=column_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = CELL_BORDER

    headers = {
        sheet.cell(row=header_row, column=i).value: i for i in range(1, n_cols + 1)
    }

    for row_index in range(header_row + 1, header_row + n_rows + 1):
        for column_index in range(1, n_cols + 1):
            sheet.cell(row=row_index, column=column_index).border = CELL_BORDER

    def apply_format(column_name: str, fmt: str, align: str = "right") -> None:
        index = headers.get(column_name)
        if not index:
            return
        for row_index in range(header_row + 1, header_row + n_rows + 1):
            cell = sheet.cell(row=row_index, column=index)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal=align)

    apply_format("Date", DATE_FORMAT, "center")
    apply_format("Retrieved At (UTC)", DATETIME_FORMAT, "center")
    apply_format("Unit", "0", "center")
    for column_name in headers:
        if column_name and (
            "Rate" in str(column_name) or str(column_name) in EXPECTED_PAIRS
        ):
            apply_format(column_name, RATE_FORMAT)

    _autosize(sheet, header_row=header_row)


def _add_excel_table(sheet, name: str, header_row: int, n_rows: int,
                     n_cols: int) -> None:
    """Register a real Excel table so filters and banding come for free."""
    if n_rows < 1:
        return
    ref = (
        f"A{header_row}:"
        f"{get_column_letter(n_cols)}{header_row + n_rows}"
    )
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    try:
        sheet.add_table(table)
    except ValueError:
        # Duplicate display name — cosmetic only, never worth failing a run.
        log.debug("Could not add table %s", name)


def _autosize(sheet, header_row: int = 1, max_width: int = 46) -> None:
    for column_index in range(1, sheet.max_column + 1):
        longest = 0
        for row_index in range(header_row, sheet.max_row + 1):
            value = sheet.cell(row=row_index, column=column_index).value
            if value is not None:
                longest = max(longest, len(str(value)))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(12, longest + 3), max_width
        )
